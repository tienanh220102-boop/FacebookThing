#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FPT University Historical Crawler (chay 1 lan)
Lay toan bo lich su bai viet tu website daihoc.fpt.edu.vn
Nguon: listing pages (tieu de, excerpt) + sitemaps (ngay dang)
Tich hop Gemini API de tom tat noi dung day du tung bai.
"""
import sys, os, json, time, re
sys.stdout.reconfigure(encoding='utf-8')

from datetime import datetime, timezone, timedelta
import xml.etree.ElementTree as ET
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

VN_TZ    = timezone(timedelta(hours=7))
HEADERS  = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
OUT_DIR  = 'Trang web FPT'
OUT_FILE = f'{OUT_DIR}/fpt_lichsu.xlsx'
JSON_OUT = f'{OUT_DIR}/fpt_lichsu.json'

# ── Cau hinh Gemini API ───────────────────────────────────────
# Nhap Gemini API Key vao day, hoac set bien moi truong GEMINI_API_KEY
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY', '')
GEMINI_MODEL   = 'gemini-2.0-flash'

LISTING_SOURCES = {
    'Tin tuc':    'https://daihoc.fpt.edu.vn/chuyen-muc/tin-tuc/page/{}/',
    'Tuyen sinh': 'https://daihoc.fpt.edu.vn/chuyen-muc/thong-bao-tuyen-sinh/page/{}/',
    'Su kien':    'https://daihoc.fpt.edu.vn/event/page/{}/',
}

# ── Buoc 1: Doc sitemaps lay URL -> Ngay ──────────────────────
def load_sitemap_dates():
    """Doc 45 sitemap XML, tra ve dict {url: date_str}."""
    url_date = {}
    ns = {'sm': 'http://www.sitemaps.org/schemas/sitemap/0.9'}
    total_sitemaps = 45

    print(f'[1/2] Doc {total_sitemaps} sitemap XML...')
    for i in range(1, total_sitemaps + 1):
        url = f'https://daihoc.fpt.edu.vn/post-sitemap{i}.xml'
        try:
            resp = requests.get(url, headers=HEADERS, timeout=15)
            if resp.status_code != 200:
                break
            root = ET.fromstring(resp.content)
            for loc in root.findall('.//sm:url', ns):
                post_url  = (loc.findtext('sm:loc', namespaces=ns) or '').strip()
                lastmod   = (loc.findtext('sm:lastmod', namespaces=ns) or '').strip()
                # Chuyen sang gio VN
                if lastmod:
                    try:
                        dt = datetime.fromisoformat(lastmod).astimezone(VN_TZ)
                        lastmod = dt.strftime('%d/%m/%Y')
                    except Exception:
                        pass
                if post_url:
                    url_date[post_url] = lastmod
        except Exception as e:
            print(f'  Loi sitemap {i}: {e}')
        print(f'  Sitemap {i:02d}/{total_sitemaps} — {len(url_date)} URLs', end='\r')
        time.sleep(0.3)

    print(f'\n  Tong cong {len(url_date)} bai trong sitemap')
    return url_date

# ── Buoc 2: Scrape listing pages lay Tieu de + Excerpt ────────
def scrape_listing(name, url_pattern):
    """Scrape tat ca trang listing cua 1 danh muc."""
    articles = []
    seen_links = set()
    page = 1
    empty_streak = 0

    while True:
        url  = url_pattern.format(page)
        try:
            resp = requests.get(url, headers=HEADERS, timeout=20)
            if resp.status_code == 404:
                break
            soup = BeautifulSoup(resp.text, 'html.parser')
        except Exception as e:
            print(f'\n  Loi trang {page}: {e}')
            if page <= 3:
                time.sleep(2)
                continue
            break

        # Tin tuc / Tuyen sinh dung class news-list-item
        # Su kien dung class all-event-card hoac event-item
        items = soup.find_all('article', class_='news-list-item')
        if not items:
            items = soup.find_all('div', class_='all-event-card')
        if not items:
            items = soup.find_all('div', class_='event-item')
        if not items:
            items = soup.find_all('article')

        added = 0
        for item in items:
            a_tag   = item.find('a', href=True)
            link    = a_tag['href'] if a_tag else ''
            title_t = (item.find(class_='news-list-title') or
                       item.find(class_='all-event-content') or
                       item.find(['h2', 'h3', 'h4']))
            title   = title_t.get_text(strip=True) if title_t else ''
            exc_t   = item.find(class_='news-list-excerpt') or item.find('p')
            excerpt = exc_t.get_text(strip=True) if exc_t else ''

            norm_link = link.rstrip('/') + '/'
            if link and title and norm_link not in seen_links:
                seen_links.add(norm_link)
                articles.append({
                    'source':  name,
                    'title':   title,
                    'link':    norm_link,
                    'excerpt': excerpt[:400],
                    'date':    '',
                })
                added += 1

        if added == 0:
            empty_streak += 1
            if empty_streak >= 2:
                break
        else:
            empty_streak = 0

        print(f'  [{name}] Trang {page} — {len(articles)} bai', end='\r')
        page += 1
        time.sleep(0.4)

    print(f'\n  [{name}] Xong: {len(articles)} bai')
    return articles

# ── Gemini: Lay full text va tom tat ─────────────────────────
def fetch_full_text(url):
    """Lay noi dung chinh day du cua bai viet."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        soup = BeautifulSoup(resp.text, 'html.parser')
        selectors = [
            '.elementor-widget-theme-post-content',
            '.entry-content',
            '.post-content',
            '.single-content',
            'article .content',
            'article',
        ]
        for sel in selectors:
            el = soup.select_one(sel)
            if el:
                for tag in el.find_all(['script', 'style', 'nav', 'footer', 'aside']):
                    tag.decompose()
                text = el.get_text(separator='\n', strip=True)
                if len(text) > 200:
                    return text[:5000]
    except Exception:
        pass
    return ''

def summarize_with_gemini(title, content):
    """Dung Gemini tom tat noi dung bai viet thanh 3-4 cau tieng Viet."""
    if not GEMINI_API_KEY or not content:
        return ''
    try:
        import google.generativeai as genai
        genai.configure(api_key=GEMINI_API_KEY)
        model = genai.GenerativeModel(GEMINI_MODEL)
        prompt = (
            'Tóm tắt bài báo sau thành 3–4 câu liền mạch bằng tiếng Việt, '
            'đủ các ý: sự kiện/hoạt động chính, số liệu & tên người/tổ chức nổi bật, '
            'kết quả hoặc ý nghĩa. Chỉ trả về đoạn tóm tắt, không thêm tiêu đề hay ghi chú.\n\n'
            f'Tiêu đề: {title}\n\nNội dung:\n{content}'
        )
        resp = model.generate_content(prompt)
        return resp.text.strip()
    except Exception as e:
        print(f'  [Gemini] Loi: {e}')
        return ''

# ── Buoc 3: Lay ngay that + full text + tom tat ───────────────
def fetch_article_details(articles):
    """Doc meta published_time + lay full content + tom tat Gemini cho tung bai."""
    total  = len(articles)
    ok_date = 0
    ok_sum  = 0
    use_gemini = bool(GEMINI_API_KEY)

    if use_gemini:
        print(f'  [Gemini] Bat: model={GEMINI_MODEL}')
    else:
        print('  [Gemini] Tat (chua co GEMINI_API_KEY — dung excerpt ngan tu listing)')

    for i, a in enumerate(articles, 1):
        try:
            resp = requests.get(a['link'], headers=HEADERS, timeout=15)
            html = resp.text

            # Lay ngay dang
            idx = html.find('article:published_time')
            if idx != -1:
                m = re.search(r'content=["\'](\d{4}-\d{2}-\d{2})', html[idx:idx+120])
                if m:
                    dt = datetime.fromisoformat(m.group(1))
                    a['date'] = dt.strftime('%d/%m/%Y')
                    ok_date += 1

            # Lay full text va tom tat neu co Gemini
            if use_gemini:
                soup = BeautifulSoup(html, 'html.parser')
                full_text = ''
                for sel in ['.elementor-widget-theme-post-content', '.entry-content',
                            '.post-content', '.single-content', 'article']:
                    el = soup.select_one(sel)
                    if el:
                        for tag in el.find_all(['script', 'style', 'nav', 'footer']):
                            tag.decompose()
                        t = el.get_text(separator='\n', strip=True)
                        if len(t) > 200:
                            full_text = t[:5000]
                            break
                if full_text:
                    summary = summarize_with_gemini(a['title'], full_text)
                    if summary:
                        a['excerpt'] = summary
                        ok_sum += 1

        except Exception:
            pass

        status = f'{ok_sum} tom tat' if use_gemini else f'{ok_date} ngay'
        print(f'  Chi tiet: {i}/{total} — {status}', end='\r')
        time.sleep(0.5 if use_gemini else 0.3)

    print(f'\n  Ket qua: {ok_date}/{total} ngay, {ok_sum}/{total} tom tat Gemini')
    return articles

# ── Helper ────────────────────────────────────────────────────
CAMPUS_MAP = [
    (['hà nội', 'ha noi', 'hanoi', 'hòa lạc', 'hoa lac'], 'Hà Nội'),
    (['tp.hcm', 'tphcm', 'tp hcm', 'hồ chí minh', 'ho chi minh'], 'TP.HCM'),
    (['cần thơ', 'can tho'], 'Cần Thơ'),
    (['đà nẵng', 'da nang'], 'Đà Nẵng'),
    (['quy nhơn', 'quy nhon'], 'Quy Nhơn'),
]

def extract_campus(title, excerpt):
    text = (title + ' ' + excerpt).lower()
    found = [label for keywords, label in CAMPUS_MAP if any(k in text for k in keywords)]
    return ', '.join(found) if found else ''

COLS = [
    'STT', 'TÊN HOẠT ĐỘNG', 'Loại sự kiện', 'NGÀY ĐĂNG', 'CAMPUS',
    'NGƯỜI/ ĐƠN VỊ, CƠ SỞ THỰC HIỆN',
    'ĐỐI TƯỢNG\n(Sinh viên, Giảng viên, Cán bộ, HS-GV bên ngoài,\nDoanh nghiệp, Cộng đồng, Người dân, Tổ chức, ...)',
    'NỘI DUNG HOẠT ĐỘNG', 'LINK ĐÃ ĐĂNG TIN',
]

COL_WIDTHS = {
    'A': 6,   # STT
    'B': 50,  # TEN HOAT DONG
    'C': 14,  # Loai su kien
    'D': 14,  # KY/NAM
    'E': 14,  # CAMPUS
    'F': 25,  # NGUOI/DON VI
    'G': 30,  # DOI TUONG
    'H': 60,  # NOI DUNG
    'I': 45,  # LINK
}

def _row(a, stt):
    return [
        stt,
        a['title'],
        a['source'],
        a['date'],
        extract_campus(a['title'], a['excerpt']),
        '',  # NGUOI/DON VI — dien tay
        '',  # DOI TUONG    — dien tay
        a['excerpt'],
        a['link'],
    ]

# ── Xuat Excel ────────────────────────────────────────────────
HDR_FILL = PatternFill(start_color='E85B2A', end_color='E85B2A', fill_type='solid')
HDR_FONT = Font(bold=True, color='FFFFFF')
TOP_FILL = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
CTR      = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

def _header(ws, cols):
    ws.append(cols)
    for cell in ws[ws.max_row]:
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = CTR

def _set_widths(ws):
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

def _autowidth(ws, max_w=60):
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 3, max_w)

def export_excel(all_articles):
    wb = openpyxl.Workbook()

    # Sheet 1: Tong quan
    ws1 = wb.active
    ws1.title = 'Tong quan'
    _header(ws1, ['Nguồn', 'Số bài', 'Bài cũ nhất', 'Bài mới nhất'])
    sources = {}
    for a in all_articles:
        sources.setdefault(a['source'], []).append(a)
    for src, arts in sources.items():
        with_date = [a for a in arts if a['date']]
        oldest = with_date[-1]['date'] if with_date else ''
        newest = with_date[0]['date']  if with_date else ''
        ws1.append([src, len(arts), oldest, newest])
    _autowidth(ws1)

    # Sheet 2: Tat ca hoat dong
    ws2 = wb.create_sheet('Tất cả hoạt động')
    _header(ws2, COLS)
    ws2.row_dimensions[1].height = 45
    _set_widths(ws2)
    for i, a in enumerate(all_articles, 1):
        ws2.append(_row(a, i))
        for cell in ws2[ws2.max_row]:
            cell.alignment = LEFT_WRAP
    _autowidth(ws2)
    _set_widths(ws2)

    # Sheet 3: Chi bai 2025 tro ve truoc
    ws3 = wb.create_sheet('Hoạt động 2025 trở về trước')
    _header(ws3, COLS)
    ws3.row_dimensions[1].height = 45
    _set_widths(ws3)
    old_arts = [a for a in all_articles if a['date'] and (
        '/2025' in a['date'] or
        (len(a['date']) >= 10 and int(a['date'].split('/')[-1][:4]) < 2025)
    )]
    for i, a in enumerate(old_arts, 1):
        ws3.append(_row(a, i))
        for cell in ws3[ws3.max_row]:
            cell.alignment = LEFT_WRAP
    _autowidth(ws3)
    _set_widths(ws3)

    wb.save(OUT_FILE)
    print(f'  Da luu: {OUT_FILE}')

# ── Main ──────────────────────────────────────────────────────
def main():
    now = datetime.now(VN_TZ)
    print(f'=== FPT Historical Crawler -- {now.strftime("%d/%m/%Y %H:%M")} ===\n')
    os.makedirs(OUT_DIR, exist_ok=True)

    # Buoc 1: Scrape listing
    print('[1/2] Scrape listing pages...')
    all_articles = []
    for name, pattern in LISTING_SOURCES.items():
        arts = scrape_listing(name, pattern)
        all_articles.extend(arts)

    # Buoc 2: Lay ngay that + full text + tom tat Gemini
    print(f'\n[2/2] Lay chi tiet cho {len(all_articles)} bai (ngay + Gemini tom tat)...')
    all_articles = fetch_article_details(all_articles)

    # Luu JSON truoc (de khong mat du lieu neu Excel bi loi)
    with open(JSON_OUT, 'w', encoding='utf-8') as f:
        json.dump({'crawled_at': now.isoformat(), 'total': len(all_articles),
                   'articles': all_articles}, f, ensure_ascii=False, indent=2)
    print(f'  Da luu JSON: {JSON_OUT}')

    # Xuat Excel
    print('\nXuat Excel...')
    export_excel(all_articles)

    print(f'\n=== Hoan thanh! Tong {len(all_articles)} bai da luu vao {OUT_DIR}/ ===')

if __name__ == '__main__':
    main()
