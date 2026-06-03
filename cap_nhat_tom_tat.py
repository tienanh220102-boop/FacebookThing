#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
cap_nhat_tom_tat.py
Cập nhật lại phần NỘI DUNG HOẠT ĐỘNG (excerpt) trong fpt_lichsu.json
bằng cách fetch full content từng bài và tóm tắt qua Gemini API.

Cách dùng:
  python cap_nhat_tom_tat.py             # Chi cập nhật bài có excerpt ngắn (<300 ký tự)
  python cap_nhat_tom_tat.py --all       # Cập nhật toàn bộ bài
  python cap_nhat_tom_tat.py --from-date 15/05/2026  # Chỉ bài từ ngày này trở đi

Yêu cầu:
  pip install requests beautifulsoup4 google-generativeai openpyxl

Cấu hình API key:
  set GEMINI_API_KEY=your_key_here   (Windows)
  hoặc sửa biến GEMINI_API_KEY bên dưới trực tiếp.
"""
import sys, os, json, time, re, argparse
sys.stdout.reconfigure(encoding='utf-8')

from datetime import datetime, timezone, timedelta
import requests
from bs4 import BeautifulSoup

# ── Cấu hình ─────────────────────────────────────────────────
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY', '')  # hoặc paste key trực tiếp vào đây
GEMINI_MODEL   = 'gemini-2.0-flash'

VN_TZ    = timezone(timedelta(hours=7))
HEADERS  = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
JSON_IN  = 'Trang web FPT/fpt_lichsu.json'
JSON_OUT = 'Trang web FPT/fpt_lichsu.json'
XLSX_OUT = 'Trang web FPT/fpt_lichsu.xlsx'


# ── Fetch full text ───────────────────────────────────────────
def fetch_full_text(url):
    """Lấy toàn bộ nội dung chính từ trang bài viết."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        soup = BeautifulSoup(resp.text, 'html.parser')
        for sel in [
            '.elementor-widget-theme-post-content',
            '.entry-content',
            '.post-content',
            '.single-content',
            'article .content',
            'article',
        ]:
            el = soup.select_one(sel)
            if el:
                for tag in el.find_all(['script', 'style', 'nav', 'footer', 'aside']):
                    tag.decompose()
                text = el.get_text(separator='\n', strip=True)
                if len(text) > 200:
                    return text[:5000]
    except Exception as e:
        print(f'    [fetch] Lỗi: {e}')
    return ''


# ── Gemini summarize ──────────────────────────────────────────
_gemini_model = None

def get_gemini_model():
    global _gemini_model
    if _gemini_model is None:
        import google.generativeai as genai
        genai.configure(api_key=GEMINI_API_KEY)
        _gemini_model = genai.GenerativeModel(GEMINI_MODEL)
    return _gemini_model

def summarize(title, content):
    """Tóm tắt bài viết bằng Gemini thành 3–4 câu tiếng Việt."""
    if not content:
        return ''
    try:
        model = get_gemini_model()
        prompt = (
            'Tóm tắt bài báo sau thành 3–4 câu liền mạch bằng tiếng Việt, '
            'đủ các ý: sự kiện/hoạt động chính, số liệu & tên người/tổ chức nổi bật (nếu có), '
            'kết quả hoặc ý nghĩa. Chỉ trả về đoạn tóm tắt, không thêm tiêu đề hay ghi chú.\n\n'
            f'Tiêu đề: {title}\n\nNội dung:\n{content}'
        )
        resp = model.generate_content(prompt)
        return resp.text.strip()
    except Exception as e:
        print(f'    [Gemini] Lỗi: {e}')
        return ''


# ── Lọc bài cần cập nhật ─────────────────────────────────────
def needs_update(article, update_all, from_date_dt):
    if update_all:
        return True

    # Lọc theo ngày nếu có --from-date
    if from_date_dt and article.get('date'):
        try:
            art_dt = datetime.strptime(article['date'], '%d/%m/%Y')
            if art_dt < from_date_dt:
                return False
        except Exception:
            pass

    excerpt = article.get('excerpt', '')
    # Cần cập nhật nếu excerpt ngắn (<300 ký tự) hoặc kết thúc bằng "…"
    return len(excerpt) < 300 or excerpt.rstrip().endswith('…')


# ── Xuất Excel ───────────────────────────────────────────────
def export_excel(articles):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        CAMPUS_MAP = [
            (['hà nội', 'ha noi', 'hanoi', 'hòa lạc', 'hoa lac'], 'Hà Nội'),
            (['tp.hcm', 'tphcm', 'tp hcm', 'hồ chí minh', 'ho chi minh'], 'TP.HCM'),
            (['cần thơ', 'can tho'], 'Cần Thơ'),
            (['đà nẵng', 'da nang'], 'Đà Nẵng'),
            (['quy nhơn', 'quy nhon'], 'Quy Nhơn'),
        ]
        def extract_campus(title, excerpt):
            text = (title + ' ' + excerpt).lower()
            found = [lb for kws, lb in CAMPUS_MAP if any(k in text for k in kws)]
            return ', '.join(found) if found else ''

        COLS = [
            'STT', 'TÊN HOẠT ĐỘNG', 'Loại sự kiện', 'NGÀY ĐĂNG', 'CAMPUS',
            'NGƯỜI/ ĐƠN VỊ, CƠ SỞ THỰC HIỆN',
            'ĐỐI TƯỢNG\n(Sinh viên, Giảng viên, Cán bộ, HS-GV bên ngoài,\nDoanh nghiệp, Cộng đồng, Người dân, Tổ chức, ...)',
            'NỘI DUNG HOẠT ĐỘNG', 'LINK ĐÃ ĐĂNG TIN',
        ]
        COL_WIDTHS = {'A':6,'B':50,'C':14,'D':14,'E':14,'F':25,'G':30,'H':60,'I':45}
        HDR_FILL = PatternFill(start_color='E85B2A', end_color='E85B2A', fill_type='solid')
        HDR_FONT = Font(bold=True, color='FFFFFF')
        CTR      = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT_TOP = Alignment(horizontal='left', vertical='top', wrap_text=True)

        wb = openpyxl.Workbook()

        # Sheet 1: Tong quan
        ws1 = wb.active; ws1.title = 'Tong quan'
        ws1.append(['Nguồn', 'Số bài', 'Bài cũ nhất', 'Bài mới nhất'])
        for cell in ws1[1]: cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = CTR
        sources = {}
        for a in articles:
            sources.setdefault(a['source'], []).append(a)
        for src, arts in sources.items():
            dated = [a for a in arts if a.get('date')]
            ws1.append([src, len(arts), dated[-1]['date'] if dated else '', dated[0]['date'] if dated else ''])

        def make_sheet(ws, rows):
            ws.append(COLS)
            ws.row_dimensions[1].height = 45
            for cell in ws[1]: cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = CTR
            for col, w in COL_WIDTHS.items(): ws.column_dimensions[col].width = w
            for i, a in enumerate(rows, 1):
                ws.append([
                    i, a['title'], a['source'], a.get('date',''),
                    extract_campus(a['title'], a.get('excerpt','')),
                    '', '', a.get('excerpt',''), a['link'],
                ])
                for cell in ws[ws.max_row]: cell.alignment = LEFT_TOP

        # Sheet 2: Tat ca
        make_sheet(wb.create_sheet('Tất cả hoạt động'), articles)

        # Sheet 3: 2025 tro ve truoc
        old = [a for a in articles if a.get('date') and (
            '/2025' in a['date'] or
            (len(a['date']) >= 10 and int(a['date'].split('/')[-1][:4]) < 2025)
        )]
        make_sheet(wb.create_sheet('Hoạt động 2025 trở về trước'), old)

        wb.save(XLSX_OUT)
        print(f'  Đã lưu Excel: {XLSX_OUT}')
    except Exception as e:
        print(f'  [Excel] Lỗi: {e}')


# ── Main ─────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='Cập nhật tóm tắt bài viết FPT bằng Gemini')
    parser.add_argument('--all', action='store_true', help='Cập nhật toàn bộ bài (mặc định: chỉ bài excerpt ngắn)')
    parser.add_argument('--from-date', metavar='DD/MM/YYYY', help='Chỉ cập nhật bài từ ngày này trở đi')
    parser.add_argument('--no-excel', action='store_true', help='Bỏ qua bước xuất Excel')
    args = parser.parse_args()

    if not GEMINI_API_KEY:
        print('LỖI: Chưa có GEMINI_API_KEY.')
        print('  Cách 1: set GEMINI_API_KEY=your_key  (trước khi chạy)')
        print('  Cách 2: Sửa biến GEMINI_API_KEY trong file này')
        sys.exit(1)

    from_date_dt = None
    if args.from_date:
        try:
            from_date_dt = datetime.strptime(args.from_date, '%d/%m/%Y')
        except ValueError:
            print('Lỗi: --from-date phải có định dạng DD/MM/YYYY'); sys.exit(1)

    # Đọc JSON
    with open(JSON_IN, encoding='utf-8') as f:
        data = json.load(f)
    articles = data['articles']
    print(f'Đã đọc {len(articles)} bài từ {JSON_IN}')

    # Lọc bài cần cập nhật
    targets = [a for a in articles if needs_update(a, args.all, from_date_dt)]
    print(f'Cần cập nhật: {len(targets)} bài\n')

    if not targets:
        print('Không có bài nào cần cập nhật.')
        return

    ok = 0
    for i, a in enumerate(targets, 1):
        print(f'[{i}/{len(targets)}] {a["title"][:60]}...')
        full_text = fetch_full_text(a['link'])
        if not full_text:
            print('    Không lấy được nội dung, bỏ qua.')
            continue
        summary = summarize(a['title'], full_text)
        if summary:
            a['excerpt'] = summary
            ok += 1
            print(f'    OK: {summary[:80]}...')
        else:
            print('    Gemini không trả về kết quả.')
        time.sleep(1)  # Tránh rate limit

    print(f'\nHoàn thành: {ok}/{len(targets)} bài đã được cập nhật tóm tắt.')

    # Lưu JSON
    data['crawled_at'] = datetime.now(VN_TZ).isoformat()
    with open(JSON_OUT, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f'Đã lưu JSON: {JSON_OUT}')

    # Xuất Excel
    if not args.no_excel:
        print('Xuất Excel...')
        export_excel(articles)

    print('\nXong!')


if __name__ == '__main__':
    main()
