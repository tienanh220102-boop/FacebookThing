#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FPT University News Crawler v2
Thu thap tin tuc tu website chinh thuc Dai hoc FPT qua RSS feed.
Chay moi ngay 1 lan luc 7h sang (Gio VN), xuat bao cao Excel.
"""
import json, os, re, time, glob
from datetime import datetime, timezone, timedelta
from email.utils import parsedate_to_datetime
import xml.etree.ElementTree as ET

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Cau hinh ──────────────────────────────────────────────────
DATA_FILE     = 'Trang web FPT/fpt_data.json'          # kho tich luy (JSON)
UNIFIED_XLSX  = 'Trang web FPT/fpt_news_tong_hop.xlsx' # 1 file Excel thong nhat (dia chi co dinh)
VN_TZ         = timezone(timedelta(hours=7))

# RSS feeds chinh thuc cua Dai hoc FPT
FEEDS = {
    'Tin tuc':    'https://daihoc.fpt.edu.vn/feed/',
    'Tuyen sinh': 'https://daihoc.fpt.edu.vn/chuyen-muc/thong-bao-tuyen-sinh/feed/',
    'Su kien':    'https://daihoc.fpt.edu.vn/event/feed/',
}

# So gio lay lui (0 = tat ca, 24 = chi lay 24h gan nhat)
HOURS_LOOKBACK = 0

# ── Lay RSS ───────────────────────────────────────────────────
def strip_html(text):
    return re.sub(r'<[^>]+>', '', text or '').strip()

def fetch_feed(name, url, since_ts=None):
    """Lay bai viet tu mot RSS feed."""
    try:
        resp = requests.get(url, timeout=15, headers={'User-Agent': 'Mozilla/5.0'})
        resp.raise_for_status()
        root = ET.fromstring(resp.content)
    except Exception as e:
        print(f'  Loi khi lay [{name}]: {e}')
        return []

    items = []
    for item in root.findall('.//item'):
        title   = strip_html(item.findtext('title') or '')
        link    = (item.findtext('link') or '').strip()
        pub_raw = (item.findtext('pubDate') or '').strip()
        desc    = strip_html(item.findtext('description') or '')[:500]

        try:
            dt_utc = parsedate_to_datetime(pub_raw)
            dt_vn  = dt_utc.astimezone(VN_TZ)
            created_display = dt_vn.strftime('%d/%m/%Y %H:%M')
            created_ts      = dt_vn.timestamp()
        except Exception:
            created_display = pub_raw
            created_ts      = 0

        if since_ts and created_ts > 0 and created_ts < since_ts:
            continue

        items.append({
            'source':     name,
            'title':      title,
            'link':       link,
            'created':    created_display,
            'created_ts': created_ts,
            'summary':    desc,
        })

    return items

# ── Helper ────────────────────────────────────────────────────
CAMPUS_MAP = [
    (['hà nội', 'ha noi', 'hanoi', 'hòa lạc', 'hoa lac'], 'Hà Nội'),
    (['tp.hcm', 'tphcm', 'tp hcm', 'hồ chí minh', 'ho chi minh'], 'TP.HCM'),
    (['cần thơ', 'can tho'], 'Cần Thơ'),
    (['đà nẵng', 'da nang'], 'Đà Nẵng'),
    (['quy nhơn', 'quy nhon'], 'Quy Nhơn'),
]

def extract_campus(title, summary):
    text = (title + ' ' + summary).lower()
    found = [label for keywords, label in CAMPUS_MAP if any(k in text for k in keywords)]
    return ', '.join(found) if found else ''

COLS = [
    'STT', 'TÊN HOẠT ĐỘNG', 'Loại sự kiện', 'NGÀY ĐĂNG', 'CAMPUS',
    'NGƯỜI/ ĐƠN VỊ, CƠ SỞ THỰC HIỆN',
    'ĐỐI TƯỢNG\n(Sinh viên, Giảng viên, Cán bộ, HS-GV bên ngoài,\nDoanh nghiệp, Cộng đồng, Người dân, Tổ chức, ...)',
    'NỘI DUNG HOẠT ĐỘNG', 'LINK ĐÃ ĐĂNG TIN',
]

COL_WIDTHS = {
    'A': 6,   # STT
    'B': 50,  # TEN HOAT DONG
    'C': 14,  # Loai su kien
    'D': 14,  # KY/NAM
    'E': 14,  # CAMPUS
    'F': 25,  # NGUOI/DON VI
    'G': 30,  # DOI TUONG
    'H': 60,  # NOI DUNG
    'I': 45,  # LINK
}

def _row(a, stt):
    return [
        stt,
        a['title'],
        a['source'],
        a['created'],
        extract_campus(a['title'], a['summary']),
        '',  # NGUOI/DON VI — dien tay
        '',  # DOI TUONG    — dien tay
        a['summary'],
        a['link'],
    ]

# ── Xuat Excel ────────────────────────────────────────────────
HDR_FILL = PatternFill(start_color='E85B2A', end_color='E85B2A', fill_type='solid')
HDR_FONT = Font(bold=True, color='FFFFFF')
TOP_FILL = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
CTR      = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

def _header(ws, cols):
    ws.append(cols)
    for cell in ws[ws.max_row]:
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = CTR

def _set_widths(ws):
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

def _autowidth(ws, max_w=60):
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 3, max_w)

def export_excel(all_articles, filename):
    """Tao file Excel 3 sheet: Tong quan / Tat ca hoat dong / 20 bai moi nhat."""
    wb = openpyxl.Workbook()
    sorted_arts = sorted(all_articles, key=lambda x: x['created_ts'], reverse=True)

    # ── Sheet 1: Tong quan nguon ──────────────────────────────
    ws1 = wb.active
    ws1.title = 'Tong quan nguon'
    _header(ws1, ['Nguồn tin', 'Số bài viết', 'Bài mới nhất'])
    sources = {}
    for a in all_articles:
        sources.setdefault(a['source'], []).append(a)
    for src, arts in sources.items():
        latest = sorted(arts, key=lambda x: x['created_ts'], reverse=True)
        ws1.append([src, len(arts), latest[0]['created'] if latest else ''])
    _autowidth(ws1)

    # ── Sheet 2: Tat ca hoat dong ─────────────────────────────
    ws2 = wb.create_sheet('Tất cả hoạt động')
    _header(ws2, COLS)
    ws2.row_dimensions[1].height = 45
    _set_widths(ws2)
    for i, a in enumerate(sorted_arts, 1):
        ws2.append(_row(a, i))
        for cell in ws2[ws2.max_row]:
            cell.alignment = LEFT_WRAP

    # ── Sheet 3: 20 bai moi nhat ──────────────────────────────
    ws3 = wb.create_sheet('20 hoạt động mới nhất')
    _header(ws3, COLS)
    ws3.row_dimensions[1].height = 45
    _set_widths(ws3)
    for i, a in enumerate(sorted_arts[:20], 1):
        ws3.append(_row(a, i))
        for cell in ws3[ws3.max_row]:
            cell.alignment = LEFT_WRAP
            if i <= 3:
                cell.fill = TOP_FILL

    wb.save(filename)
    print(f'  Xuat thanh cong: {filename}')

# ── Kho tich luy (dedupe theo link) ───────────────────────────
def _norm_link(link):
    return (link or '').rstrip('/') + '/'

def load_store():
    """Doc kho bai viet da tich luy tu cac lan chay truoc."""
    try:
        with open(DATA_FILE, encoding='utf-8') as f:
            return json.load(f).get('articles', [])
    except (FileNotFoundError, json.JSONDecodeError):
        return []

def merge_articles(existing, new):
    """Gop bai moi vao kho, dedupe theo link chuan hoa; bo sung created_ts neu ban cu thieu."""
    by_link = {_norm_link(a['link']): a for a in existing}
    for a in new:
        key = _norm_link(a['link'])
        if key not in by_link:
            by_link[key] = a
        elif a.get('created_ts', 0) and not by_link[key].get('created_ts'):
            by_link[key] = a
    return list(by_link.values())

def prune_snapshots(keep_file):
    """Xoa cac snapshot ngay cu, chi giu lai snapshot moi nhat (keep_file).
       File tong hop (UNIFIED_XLSX) luon duoc giu."""
    keep = {os.path.basename(keep_file), os.path.basename(UNIFIED_XLSX)}
    removed = 0
    for path in glob.glob('Trang web FPT/fpt_news_*.xlsx'):
        if os.path.basename(path) not in keep:
            try:
                os.remove(path)
                removed += 1
            except OSError:
                pass
    if removed:
        print(f'  Da don {removed} snapshot cu')

# ── Main ──────────────────────────────────────────────────────
def main():
    now_vn = datetime.now(VN_TZ)
    if HOURS_LOOKBACK > 0:
        since_ts = (now_vn - timedelta(hours=HOURS_LOOKBACK)).timestamp()
        lookback_text = f'{HOURS_LOOKBACK}h gan nhat'
    else:
        since_ts = None
        lookback_text = 'tat ca bai (khong gioi han thoi gian)'

    print(f'=== FPT University Crawler v2 -- {now_vn.strftime("%d/%m/%Y %H:%M")} (Gio VN) ===')
    print(f'Thu thap {lookback_text}...')

    all_articles = []

    for name, url in FEEDS.items():
        print(f'\n[{name}] Dang lay du lieu...')
        articles = fetch_feed(name, url, since_ts)
        print(f'  Lay duoc {len(articles)} bai')
        all_articles.extend(articles)
        time.sleep(0.5)

    if not all_articles:
        print('\nKhong co bai viet nao. Kiem tra ket noi mang hoac RSS feed.')
        return

    os.makedirs('Trang web FPT', exist_ok=True)

    # 1) Snapshot theo ngay — chi cac bai lay duoc trong lan chay nay
    date_str      = now_vn.strftime('%Y%m%d')
    snapshot_file = f'Trang web FPT/fpt_news_{date_str}.xlsx'
    print(f'\nXuat snapshot ngay...')
    export_excel(all_articles, snapshot_file)
    prune_snapshots(snapshot_file)   # chi giu snapshot moi nhat + file tong hop

    # 2) Gop vao kho tich luy -> 1 file Excel + JSON thong nhat (dia chi co dinh)
    existing = load_store()
    store    = merge_articles(existing, all_articles)
    added    = len(store) - len(existing)
    print(f'\nGop kho: +{added} bai moi, tong cong {len(store)} bai')
    export_excel(store, UNIFIED_XLSX)

    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump({
            'crawled_at':   now_vn.isoformat(),
            'total':        len(store),
            'articles':     store,
        }, f, ensure_ascii=False, indent=2)

    print(f'\n=== Hoan thanh. +{added} bai moi / {len(store)} bai trong kho ===')
    print(f'  File thong nhat: {UNIFIED_XLSX}')
    print(f'  Du lieu JSON:    {DATA_FILE}')

if __name__ == '__main__':
    main()
