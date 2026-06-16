#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Dong bo fpt_lichsu.json -> fpt_lichsu.xlsx"""
import sys, json
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

JSON_IN  = 'Trang web FPT/fpt_lichsu.json'
XLSX_OUT = 'Trang web FPT/fpt_lichsu_new.xlsx'

CAMPUS_MAP = [
    (['hà nội','ha noi','hanoi','hòa lạc','hoa lac'], 'Hà Nội'),
    (['tp.hcm','tphcm','tp hcm','hồ chí minh','ho chi minh'], 'TP.HCM'),
    (['cần thơ','can tho'], 'Cần Thơ'),
    (['đà nẵng','da nang'], 'Đà Nẵng'),
    (['quy nhơn','quy nhon'], 'Quy Nhơn'),
]
def campus(title, exc):
    text = (title+' '+exc).lower()
    found = [lb for kws,lb in CAMPUS_MAP if any(k in text for k in kws)]
    return ', '.join(found)

COLS = [
    'STT','TÊN HOẠT ĐỘNG','Loại sự kiện','NGÀY ĐĂNG','CAMPUS',
    'NGƯỜI/ ĐƠN VỊ, CƠ SỞ THỰC HIỆN',
    'ĐỐI TƯỢNG\n(Sinh viên, Giảng viên, Cán bộ, HS-GV bên ngoài,\nDoanh nghiệp, Cộng đồng, Người dân, Tổ chức, ...)',
    'NỘI DUNG HOẠT ĐỘNG','LINK ĐÃ ĐĂNG TIN',
]
COL_W = {'A':6,'B':50,'C':14,'D':14,'E':14,'F':25,'G':30,'H':60,'I':45}
HDR_FILL = PatternFill(start_color='E85B2A',end_color='E85B2A',fill_type='solid')
HDR_FONT = Font(bold=True,color='FFFFFF')
CTR      = Alignment(horizontal='center',vertical='center',wrap_text=True)
LT       = Alignment(horizontal='left',vertical='top',wrap_text=True)

def make_sheet(ws, rows):
    ws.append(COLS)
    ws.row_dimensions[1].height = 45
    for c in ws[1]: c.font=HDR_FONT; c.fill=HDR_FILL; c.alignment=CTR
    for col,w in COL_W.items(): ws.column_dimensions[col].width = w
    for i,a in enumerate(rows,1):
        ws.append([
            i, a['title'], a['source'], a.get('date',''),
            campus(a['title'], a.get('excerpt','')),
            '', '', a.get('excerpt',''), a['link'],
        ])
        for c in ws[ws.max_row]: c.alignment = LT

with open(JSON_IN, encoding='utf-8') as f:
    data = json.load(f)
arts = data['articles']
print(f'Loaded {len(arts)} articles')

wb = openpyxl.Workbook()

# Sheet 1: Tong quan
ws1 = wb.active; ws1.title = 'Tong quan'
ws1.append(['Nguồn','Số bài','Bài cũ nhất','Bài mới nhất'])
for c in ws1[1]: c.font=HDR_FONT; c.fill=HDR_FILL; c.alignment=CTR
src_map = {}
for a in arts: src_map.setdefault(a['source'],[]).append(a)
for src,lst in src_map.items():
    dated = [a for a in lst if a.get('date')]
    ws1.append([src, len(lst), dated[-1]['date'] if dated else '', dated[0]['date'] if dated else ''])

# Sheet 2: Tat ca
make_sheet(wb.create_sheet('Tất cả hoạt động'), arts)

# Sheet 3: 2025 tro ve truoc
old = [a for a in arts if a.get('date') and (
    '/2025' in a['date'] or
    (len(a['date'])>=10 and int(a['date'].split('/')[-1][:4])<2025)
)]
make_sheet(wb.create_sheet('Hoạt động 2025 trở về trước'), old)

wb.save(XLSX_OUT)
print(f'Saved: {XLSX_OUT}  ({len(arts)} rows)')
