#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Facebook Fanpage Crawler v1
Thu thap du lieu tu cac fanpage bao chi public qua Graph API chinh thuc.
Chay moi ngay 1 lan luc 7h sang (Gio VN), xuat bao cao Excel.
"""
import json, os, time
from datetime import datetime, timezone, timedelta

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Cau hinh ──────────────────────────────────────────────────
# Lay FB_TOKEN tai: developers.facebook.com → Tools → Graph API Explorer
# Format token: <App-ID>|<App-Secret>  (App Access Token)
FB_TOKEN  = os.environ.get('FB_TOKEN', '')
GRAPH_URL = 'https://graph.facebook.com/v19.0'
DATA_FILE = 'data/fb_data.json'
VN_TZ     = timezone(timedelta(hours=7))

# Danh sach fanpage can theo doi (username hoac Page ID)
# Them/bot bang cach chinh sua dict nay
PAGES = {
    'FPT University HCM': 'FPTU.HCM',
}

# Lay du lieu bao nhieu gio gan nhat
# 0 = lay tat ca bai (khong gioi han) | 24 = chi lay 24h gan nhat
HOURS_LOOKBACK = 0

# ── Goi API ───────────────────────────────────────────────────
def api_get(endpoint, params=None):
    """Goi Graph API, tra ve dict. None neu co loi."""
    p = dict(params or {})
    p['access_token'] = FB_TOKEN
    try:
        resp = requests.get(f'{GRAPH_URL}/{endpoint}', params=p, timeout=15)
        data = resp.json()
        if 'error' in data:
            print(f'  API Error: {data["error"].get("message", data["error"])}')
            return None
        return data
    except Exception as e:
        print(f'  Request loi: {e}')
        return None

def get_page_info(page_id):
    """Lay thong tin tong quan cua Page."""
    return api_get(page_id, {
        'fields': 'id,name,fan_count,followers_count,about,category,verification_status'
    })

def get_posts(page_id, since_ts=None):
    """Lay tat ca posts cua Page tu moc thoi gian since_ts (Unix timestamp)."""
    params = {
        'fields': (
            'id,message,story,created_time,full_picture,'
            'reactions.summary(true),'
            'comments.summary(true),'
            'shares'
        ),
        'limit': 100,
        'access_token': FB_TOKEN,
    }
    if since_ts:
        params['since'] = int(since_ts)

    all_posts = []
    url = f'{GRAPH_URL}/{page_id}/posts'
    current_params = params

    while url:
        try:
            resp = requests.get(url, params=current_params, timeout=15)
            data = resp.json()
        except Exception as e:
            print(f'  Request loi: {e}')
            break
        if 'error' in data:
            print(f'  API Error: {data["error"].get("message")}')
            break
        all_posts.extend(data.get('data', []))
        next_url = data.get('paging', {}).get('next')
        # next_url da chua san access_token, khong can them params
        url = next_url
        current_params = {}
        time.sleep(0.5)

    return all_posts

# ── Xu ly du lieu ─────────────────────────────────────────────
def parse_post(post, page_name):
    """Chuan hoa 1 post thanh dict don gian de xuat Excel."""
    created_raw = post.get('created_time', '')
    try:
        dt_utc = datetime.fromisoformat(created_raw.replace('Z', '+00:00'))
        dt_vn  = dt_utc.astimezone(VN_TZ)
        created_display = dt_vn.strftime('%d/%m/%Y %H:%M')
        created_ts      = dt_vn.timestamp()
    except Exception:
        created_display = created_raw
        created_ts      = 0

    reactions  = post.get('reactions', {}).get('summary', {}).get('total_count', 0)
    comments   = post.get('comments',  {}).get('summary', {}).get('total_count', 0)
    shares     = post.get('shares',    {}).get('count', 0)
    engagement = reactions + comments + shares
    content    = (post.get('message') or post.get('story') or '')[:500]

    # Chuyen post_id "page_postnum" thanh link
    raw_id = post.get('id', '')
    parts  = raw_id.split('_')
    link   = f'https://facebook.com/{parts[0]}/posts/{parts[1]}' if len(parts) == 2 else ''

    return {
        'page':        page_name,
        'post_id':     raw_id,
        'created':     created_display,
        'created_ts':  created_ts,
        'content':     content,
        'reactions':   reactions,
        'comments':    comments,
        'shares':      shares,
        'engagement':  engagement,
        'image':       post.get('full_picture', ''),
        'link':        link,
    }

# ── Xuat Excel ────────────────────────────────────────────────
HDR_FILL = PatternFill(start_color='1877F2', end_color='1877F2', fill_type='solid')
HDR_FONT = Font(bold=True, color='FFFFFF')
TOP_FILL = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
CTR      = Alignment(horizontal='center', vertical='center', wrap_text=True)

def _header(ws, cols):
    ws.append(cols)
    for cell in ws[ws.max_row]:
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = CTR

def _autowidth(ws, max_w=55):
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 3, max_w)

def export_excel(pages_info, all_posts, filename):
    """Tao file Excel 3 sheet: Tong quan / Tat ca Posts / Top 20 Viral."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Tong quan Page ───────────────────────────────
    ws1 = wb.active
    ws1.title = 'Tong quan Page'
    _header(ws1, [
        'Trang bao', 'Followers', 'Bai hom nay',
        'Tong Reactions', 'Tong Comments', 'Tong Shares',
        'Avg Engagement/bai', 'Danh muc', 'Verified'
    ])
    for page_name, info in pages_info.items():
        pp   = [p for p in all_posts if p['page'] == page_name]
        t_r  = sum(p['reactions'] for p in pp)
        t_c  = sum(p['comments']  for p in pp)
        t_s  = sum(p['shares']    for p in pp)
        avg  = round((t_r + t_c + t_s) / max(len(pp), 1), 1)
        ws1.append([
            page_name,
            info.get('fan_count', 0) if info else 0,
            len(pp), t_r, t_c, t_s, avg,
            info.get('category', '') if info else '',
            'Co' if (info or {}).get('verification_status') == 'blue_verified' else 'Khong',
        ])
    _autowidth(ws1)

    # ── Sheet 2: Tat ca Posts ─────────────────────────────────
    ws2 = wb.create_sheet('Tat ca Posts')
    _header(ws2, [
        'Trang bao', 'Thoi gian', 'Noi dung bai viet',
        'Reactions', 'Comments', 'Shares', 'Tong Engagement', 'Link'
    ])
    ws2.column_dimensions['C'].width = 65
    for p in sorted(all_posts, key=lambda x: x['created_ts'], reverse=True):
        ws2.append([
            p['page'], p['created'], p['content'],
            p['reactions'], p['comments'], p['shares'], p['engagement'], p['link']
        ])
    _autowidth(ws2)

    # ── Sheet 3: Top 20 Viral ─────────────────────────────────
    ws3 = wb.create_sheet('Top 20 Viral')
    _header(ws3, [
        'Hang', 'Trang bao', 'Thoi gian', 'Noi dung bai viet',
        'Reactions', 'Comments', 'Shares', 'Tong Engagement', 'Link'
    ])
    ws3.column_dimensions['D'].width = 65
    top20 = sorted(all_posts, key=lambda x: x['engagement'], reverse=True)[:20]
    for rank, p in enumerate(top20, 1):
        ws3.append([
            rank, p['page'], p['created'], p['content'],
            p['reactions'], p['comments'], p['shares'], p['engagement'], p['link']
        ])
        if rank <= 3:
            for cell in ws3[ws3.max_row]:
                cell.fill = TOP_FILL
    _autowidth(ws3)

    wb.save(filename)
    print(f'  Xuat thanh cong: {filename}')

# ── Main ──────────────────────────────────────────────────────
def main():
    if not FB_TOKEN:
        print('CHUA CO TOKEN!')
        print('Huong dan: vao developers.facebook.com → Tools → Graph API Explorer')
        print('           Chon app → Generate Access Token → copy vao bien FB_TOKEN')
        return

    now_vn   = datetime.now(VN_TZ)
    if HOURS_LOOKBACK > 0:
        since_ts = (now_vn - timedelta(hours=HOURS_LOOKBACK)).timestamp()
        since_display = datetime.fromtimestamp(since_ts, tz=VN_TZ).strftime('%d/%m %H:%M')
        lookback_text = f'tu {since_display} den bay gio ({HOURS_LOOKBACK}h)'
    else:
        since_ts = None
        lookback_text = 'tat ca bai dang (khong gioi han thoi gian)'

    print(f'=== Facebook Crawler v1 — {now_vn.strftime("%d/%m/%Y %H:%M")} (Gio VN) ===')
    print(f'Thu thap {lookback_text}...')

    pages_info = {}
    all_posts  = []

    for page_name, page_id in PAGES.items():
        print(f'\n[{page_name}] Dang xu ly...')

        info = get_page_info(page_id)
        pages_info[page_name] = info
        if info:
            print(f'  Followers: {info.get("fan_count", 0):,}')
        else:
            print('  Khong lay duoc thong tin page')

        posts_raw = get_posts(page_id, since_ts=since_ts if HOURS_LOOKBACK > 0 else None)
        print(f'  Lay duoc {len(posts_raw)} bai dang')

        for p in posts_raw:
            all_posts.append(parse_post(p, page_name))

        time.sleep(1)

    if not all_posts:
        print('\nKhong co bai dang nao. Kiem tra lai token hoac danh sach page.')
        return

    # Xuat Excel
    os.makedirs('data', exist_ok=True)
    date_str   = now_vn.strftime('%Y%m%d')
    excel_file = f'data/fb_baochi_{date_str}.xlsx'
    print(f'\nXuat Excel...')
    export_excel(pages_info, all_posts, excel_file)

    # Luu raw JSON
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump({
            'crawled_at': now_vn.isoformat(),
            'total_posts': len(all_posts),
            'posts': all_posts,
        }, f, ensure_ascii=False, indent=2)

    print(f'\n=== Hoan thanh. {len(all_posts)} bai tu {len(pages_info)} trang bao ===')

if __name__ == '__main__':
    main()
